import pandas as pd
from difflib import SequenceMatcher
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font
import re

def get_table_boundaries():
    print("Enter the starting row of the table:")
    start_row = int(input())
    print("Enter the ending row of the table:")
    end_row = int(input())
    print("Enter the starting column of the table  (A=1, B=2, C=3 ...):")
    start_col = int(input())
    print("Enter the ending column of the table  (A=1, B=2, C=3 ...):")
    end_col = int(input())
    return start_row, end_row, start_col, end_col

def clean_text(text):
    if isinstance(text, str):
        # Replace non-breaking spaces, strip, and collapse multiple spaces
        text = text.replace('\xa0', ' ').strip()
        text = re.sub(r'\s+', ' ', text)
        return text
    return text

def build_commite_douar_dict():
    start_row, end_row, start_col, end_col = get_table_boundaries()
    df = pd.read_excel('Douars.xlsx', header=9)
    table = df.iloc[start_row-1:end_row, start_col-1:end_col]
    commite_dict = {}
    for _, row in table.iterrows():
        commite = clean_text(row.iloc[0])  # First column (Commune)
        douar = clean_text(row.iloc[1])    # Second column (Douar)
        if pd.isnull(commite) or pd.isnull(douar):
            continue
        commite_dict.setdefault(commite, []).append(douar)
    # Sort douars in each committee
    for commite in commite_dict:
        commite_dict[commite] = sorted(commite_dict[commite], key=lambda x: x.lower())
    # Print in a readable format
    for commite, douars in commite_dict.items():
        print(f"Committee: {commite}")
        for douar in douars:
            print(f"  - {douar}")
        print()  # Empty line between committees
    return commite_dict

def group_similar_items(items, threshold=0.8):
    # Build similarity graph using names with all spaces removed for comparison
    n = len(items)
    adj = [[] for _ in range(n)]
    def norm(x):
        return x.replace(' ', '').lower()
    for i in range(n):
        for j in range(i+1, n):
            ratio = SequenceMatcher(None, norm(items[i]), norm(items[j])).ratio()
            if ratio >= threshold:
                adj[i].append(j)
                adj[j].append(i)
    # Find connected components (groups)
    visited = [False]*n
    groups = []
    for i in range(n):
        if not visited[i]:
            stack = [i]
            group = []
            while stack:
                node = stack.pop()
                if not visited[node]:
                    visited[node] = True
                    group.append(items[node])
                    stack.extend(adj[node])
            if len(group) > 1:
                groups.append(group)
    return groups

def merge_similar_committees(commite_dict, threshold=0.8):
    commitees = list(commite_dict.keys())
    groups = group_similar_items(commitees, threshold)
    merged = set()
    for group in groups:
        print(f"\nThese committees are similar (>{int(threshold*100)}%):")
        for idx, name in enumerate(group, 1):
            print(f"{idx}. {name}")
        answer = input("Are these the same committee? (y/n): ").strip().lower()
        if answer == 'y':
            print("Which name do you want to keep? Enter the number:")
            for idx, name in enumerate(group, 1):
                print(f"{idx}. {name}")
            choice = int(input("Your choice: ").strip())
            keep = group[choice-1]
            for name in group:
                if name != keep:
                    commite_dict[keep].extend(commite_dict[name])
                    merged.add(name)
    # Remove merged committees
    for comm in merged:
        del commite_dict[comm]
    # Sort douars in each committee after merging
    for commite in commite_dict:
        commite_dict[commite] = sorted(commite_dict[commite], key=lambda x: x.lower())
    print("\nMerged committee dictionary:")
    for commite, douars in commite_dict.items():
        print(f"Committee: {commite}")
        for douar in douars:
            print(f"  - {douar}")
        print()
    return commite_dict

def find_similar_groups(items, threshold=0.8):
    groups = []
    used = set()
    for i, item1 in enumerate(items):
        if i in used:
            continue
        group = [item1]
        for j, item2 in enumerate(items):
            if i == j or j in used:
                continue
            ratio = SequenceMatcher(None, item1.lower(), item2.lower()).ratio()
            if ratio >= threshold:
                group.append(item2)
        if len(group) > 1:
            groups.append(group)
            used.update([items.index(g) for g in group])
    return groups

def group_similar_items(items, threshold=0.8):
    # Build similarity graph
    n = len(items)
    adj = [[] for _ in range(n)]
    for i in range(n):
        for j in range(i+1, n):
            ratio = SequenceMatcher(None, items[i].lower(), items[j].lower()).ratio()
            if ratio >= threshold:
                adj[i].append(j)
                adj[j].append(i)
    # Find connected components (groups)
    visited = [False]*n
    groups = []
    for i in range(n):
        if not visited[i]:
            stack = [i]
            group = []
            while stack:
                node = stack.pop()
                if not visited[node]:
                    visited[node] = True
                    group.append(items[node])
                    stack.extend(adj[node])
            if len(group) > 1:
                groups.append(group)
    return groups

def merge_similar_douars(commite_dict, threshold=0.8):
    for commite, douars in commite_dict.items():
        groups = group_similar_items(douars, threshold)
        for group in groups:
            print(f"\nIn commune '{commite}', these douars are similar (>{int(threshold*100)}%):")
            for idx, name in enumerate(group, 1):
                print(f"{idx}. {name}")
            answer = input("Are these the same douar? (y/n): ").strip().lower()
            if answer == 'y':
                print("Which name do you want to keep? Enter the number:")
                for idx, name in enumerate(group, 1):
                    print(f"{idx}. {name}")
                choice = int(input("Your choice: ").strip())
                keep = group[choice-1]
                # Remove all group members from douars, then add the chosen one
                commite_dict[commite] = [d for d in douars if d not in group] + [keep]
                douars = commite_dict[commite]  # update for next group
    print("\nFinal commune dictionary after douar merging:")
    for commite, douars in commite_dict.items():
        print(f"Commune: {commite}")
        for douar in douars:
            print(f"  - {douar}")
        print()
    return commite_dict

def find_similar_douars(commite_dict, threshold=0.8):
    from difflib import SequenceMatcher
    similar_douars_dict = {}
    for commite, douars in commite_dict.items():
        douar_groups = []
        used = set()
        for i, d1 in enumerate(douars):
            if i in used:
                continue
            group = [d1]
            for j, d2 in enumerate(douars):
                if i == j or j in used:
                    continue
                ratio = SequenceMatcher(None, d1.lower(), d2.lower()).ratio()
                if ratio >= threshold:
                    group.append(d2)
                    used.add(j)
            if len(group) > 1:
                douar_groups.append(group)
                used.add(i)
        if douar_groups:
            similar_douars_dict[commite] = douar_groups

    # Print the results
    for commite, groups in similar_douars_dict.items():
        print(f"\nCommune: {commite}")
        for group in groups:
            print("  Similar douars group:")
            for douar in group:
                print(f"    - {douar}")
    return similar_douars_dict

def export_to_excel(commite_dict, similar_douars_dict, filename='committees_output.xlsx'):
    # Prepare data for cleaned communes
    cleaned_rows = []
    for commune, douars in commite_dict.items():
        for douar in douars:
            cleaned_rows.append({'Commune': commune, 'Douar': douar})
    cleaned_df = pd.DataFrame(cleaned_rows)

    # Prepare data for probable duplicates
    duplicate_rows = []
    for commune, groups in similar_douars_dict.items():
        for group in groups:
            duplicate_rows.append({
                'Commune': commune,
                'Probable Duplicate Douars': ', '.join(group)
            })
    duplicates_df = pd.DataFrame(duplicate_rows)

    # Write both sheets
    with pd.ExcelWriter(filename, engine='openpyxl') as writer:
        cleaned_df.to_excel(writer, sheet_name='Cleaned Communes', index=False)
        duplicates_df.to_excel(writer, sheet_name='Probable Duplicates', index=False)

    # Format the Excel file
    wb = load_workbook(filename)
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        # Center all text, make header big and bold
        for row in ws.iter_rows():
            for cell in row:
                cell.alignment = Alignment(horizontal='center', vertical='center')
                if cell.row == 1:
                    cell.font = Font(size=14, bold=True)
        # Merge same commune names and bold them
        if ws.max_column >= 1:
            col_letter = 'A'
            prev_value = None
            start_row = 2
            for row in range(2, ws.max_row + 2):  # +2 to handle last group
                value = ws[f'{col_letter}{row}'].value if row <= ws.max_row else None
                if value != prev_value and prev_value is not None:
                    if row - start_row > 1:
                        ws.merge_cells(f'{col_letter}{start_row}:{col_letter}{row-1}')
                        ws[f'{col_letter}{start_row}'].font = Font(bold=True)
                    else:
                        ws[f'{col_letter}{start_row}'].font = Font(bold=True)
                    start_row = row
                prev_value = value
    wb.save(filename)
    print(f"Excel file '{filename}' created and formatted.")

# Example usage after building and merging committees:
# commite_dict = build_commite_douar_dict()
# merge_similar_committees(commite_dict)
# similar_douars_dict = collect_similar_douars(commite_dict)
# print(similar_douars_dict)

if __name__ == "__main__":
    commite_dict = build_commite_douar_dict()
    merge_similar_committees(commite_dict)
    # Find similar douars in each commune and append to a new dictionary
    similar_douars_dict = {}
    for commite, douars in commite_dict.items():
        groups = group_similar_items(douars, threshold=0.8)
        if groups:
            similar_douars_dict[commite] = [sorted(group, key=lambda x: x.lower()) for group in groups]
    print("\nSimilar douars in each commune:")
    print(similar_douars_dict)

    # After building similar_douars_dict
    for commite, groups in similar_douars_dict.items():
        for group in groups:
            # Keep only the first douar in the group, remove the rest from commite_dict
            keep = group[0]
            for douar in group[1:]:
                if douar in commite_dict[commite]:
                    commite_dict[commite].remove(douar)

    # Remove all douars that are in any similar group from the original commite_dict
    for commite, groups in similar_douars_dict.items():
        to_remove = set()
        for group in groups:
            to_remove.update(group)
        commite_dict[commite] = [d for d in commite_dict[commite] if d not in to_remove]

    print("\nCleaned commune dictionary (duplicates removed):")
    for commite, douars in commite_dict.items():
        print(f"Commune: {commite}")
        for douar in douars:
            print(f"  - {douar}")
        print()

    # Export results to Excel
    export_to_excel(commite_dict, similar_douars_dict)



